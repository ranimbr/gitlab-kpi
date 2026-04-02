"""
api/routers/export.py

Export PDF et Excel des rapports KPI.

Endpoints :
    GET /export/kpis/excel  → fichier .xlsx téléchargeable
    GET /export/kpis/pdf    → fichier .pdf téléchargeable

Dépendances à installer :
    pip install openpyxl reportlab
"""
import io
import logging
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.api.dependencies import get_current_user
from app.database.session import get_db
from app.models.app_user import AppUser
from app.models.kpi_snapshot import KpiSnapshot
from app.repositories.kpi_snapshot_repository import KpiSnapshotRepository
from app.repositories.period_repository import PeriodRepository
from app.repositories.site_repository import SiteRepository

logger        = logging.getLogger(__name__)
router        = APIRouter(prefix="/export", tags=["Export"])
snapshot_repo = KpiSnapshotRepository()
period_repo   = PeriodRepository()
site_repo     = SiteRepository()

MOIS_FR = {
    1: "Janvier",   2: "Février",   3: "Mars",      4: "Avril",
    5: "Mai",       6: "Juin",      7: "Juillet",   8: "Août",
    9: "Septembre", 10: "Octobre",  11: "Novembre", 12: "Décembre",
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _resolve_period_and_snapshots(
    db:         Session,
    project_id: int,
    period_id:  Optional[int],
    site_id:    Optional[int],
):
    """
    Résout la période et récupère les snapshots filtrés.
    Mutualisé entre les deux endpoints export.
    """
    if period_id is None:
        snap = (
            db.query(KpiSnapshot)
            .filter(KpiSnapshot.project_id == project_id)
            .order_by(KpiSnapshot.snapshot_date.desc())
            .first()
        )
        if not snap:
            raise HTTPException(
                status_code=404,
                detail=f"Aucun snapshot trouvé pour le projet {project_id}.",
            )
        period_id = snap.period_id

    period = period_repo.get_by_id(db, period_id)
    if not period:
        raise HTTPException(status_code=404, detail="Période introuvable.")

    snapshots = snapshot_repo.get_all_by_period(db, period_id)

    # Filtrer par site si demandé, sinon garder uniquement les snapshots
    # de niveau site (developer_id IS NULL) pour éviter les doublons dans l'export
    if site_id is not None:
        snapshots = [s for s in snapshots if s.site_id == site_id and s.developer_id is None]
    else:
        snapshots = [s for s in snapshots if s.developer_id is None]

    return period, snapshots


def _get_site_name(db: Session, site_id: Optional[int]) -> str:
    """Retourne le nom du site ou 'Global' si site_id est None."""
    if site_id is None:
        return "Global"
    site_obj = site_repo.get_by_id(db, site_id)
    return site_obj.name if site_obj else f"Site {site_id}"


def _format_rate(value: Optional[float]) -> str:
    if value is None:
        return "N/A"
    return f"{value * 100:.1f}%"


def _format_float(value: Optional[float], decimals: int = 2) -> str:
    if value is None:
        return "N/A"
    return f"{value:.{decimals}f}"


# ── Excel Export ──────────────────────────────────────────────────────────────

@router.get(
    "/kpis/excel",
    summary="Export Excel — rapport KPI par projet/période/site",
    response_description="Fichier .xlsx téléchargeable",
)
def export_kpis_excel(
    project_id: int           = Query(..., description="ID du projet GitLab"),
    period_id:  Optional[int] = Query(default=None, description="ID de la période (défaut: dernière)"),
    site_id:    Optional[int] = Query(default=None, description="Filtrer par site (défaut: tous)"),
    db:         Session       = Depends(get_db),
    _:          AppUser       = Depends(get_current_user),
):
    """
    Génère un fichier Excel (.xlsx) avec le rapport KPI complet.

    Colonnes : Site | Développeurs | MR Rate/Site | Approved MR Rate |
               Merged MR Rate | Commit Rate/Site | NB Commits/Projet |
               Avg Review Time | Developer Score

    Feuille supplémentaire : résumé des deltas (comparaison vs période précédente).
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl non installé. Exécutez : pip install openpyxl",
        )

    period, snapshots = _resolve_period_and_snapshots(db, project_id, period_id, site_id)
    period_label      = f"{MOIS_FR.get(period.month, '')} {period.year}"

    wb = openpyxl.Workbook()

    # ── Feuille 1 : KPIs ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = f"KPIs {period.month:02d}-{period.year}"

    BLUE   = "2563EB"
    WHITE  = "FFFFFF"
    LIGHT  = "EFF6FF"

    header_fill = PatternFill("solid", fgColor=BLUE)
    header_font = Font(bold=True, color=WHITE, size=11)
    title_font  = Font(bold=True, size=13, color=BLUE)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    # Titre
    ws.merge_cells("A1:I1")
    title_cell       = ws["A1"]
    title_cell.value = f"Rapport KPI GitLab — {period_label} — Projet #{project_id}"
    title_cell.font  = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # En-têtes colonnes
    headers = [
        "Site",
        "Développeurs",
        "MR Rate / Site",
        "Approved MR Rate",
        "Merged MR Rate",
        "Commit Rate / Site",
        "NB Commits / Projet",
        "Avg Review Time (h)",
        "Developer Score",
    ]
    for col_idx, header in enumerate(headers, 1):
        cell           = ws.cell(row=2, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border
    ws.row_dimensions[2].height = 35

    # Données
    alt_fill = PatternFill("solid", fgColor=LIGHT)
    for row_idx, snap in enumerate(snapshots, 3):
        row_fill = alt_fill if row_idx % 2 == 0 else None
        values   = [
            _get_site_name(db, snap.site_id),
            snap.nb_developers,
            _format_float(snap.mr_rate_per_site),
            _format_rate(snap.approved_mr_rate),
            _format_rate(snap.merged_mr_rate),
            _format_float(snap.commit_rate_per_site),
            snap.nb_commits_per_project,
            _format_float(snap.avg_review_time_hours, 1),
            _format_float(snap.developer_score),
        ]
        for col_idx, value in enumerate(values, 1):
            cell           = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = thin_border
            if row_fill:
                cell.fill = row_fill

    # Largeurs colonnes
    col_widths = [20, 14, 16, 18, 16, 18, 20, 20, 16]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Feuille 2 : Deltas (comparaison vs période précédente) ───────────────
    ws2        = wb.create_sheet(title="Évolution")
    ws2["A1"]  = f"Évolution vs période précédente — {period_label}"
    ws2["A1"].font      = title_font
    ws2["A1"].alignment = Alignment(horizontal="center")
    ws2.merge_cells("A1:G1")
    ws2.row_dimensions[1].height = 28

    delta_headers = [
        "Site", "Δ MR Rate", "Δ Approved MR Rate",
        "Δ Merged MR Rate", "Δ Commit Rate",
        "Δ Commits/Projet", "Δ Avg Review Time",
    ]
    for col_idx, h in enumerate(delta_headers, 1):
        cell           = ws2.cell(row=2, column=col_idx, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border    = thin_border

    green_fill = PatternFill("solid", fgColor="D1FAE5")
    red_fill   = PatternFill("solid", fgColor="FEE2E2")

    for row_idx, snap in enumerate(snapshots, 3):
        delta_values = [
            _get_site_name(db, snap.site_id),
            _format_float(snap.delta_mr_rate),
            _format_rate(snap.delta_approved_mr_rate) if snap.delta_approved_mr_rate is not None else "N/A",
            _format_rate(snap.delta_merged_mr_rate)   if snap.delta_merged_mr_rate   is not None else "N/A",
            _format_float(snap.delta_commit_rate),
            snap.delta_nb_commits if snap.delta_nb_commits is not None else "N/A",
            _format_float(snap.delta_avg_review_time, 1),
        ]
        for col_idx, value in enumerate(delta_values, 1):
            cell           = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center")
            cell.border    = thin_border
            # Coloration automatique : vert si amélioration, rouge si régression
            if col_idx > 1 and value not in ("N/A", None):
                try:
                    num = float(str(value).replace("%", ""))
                    # Pour review_time, une valeur négative = amélioration
                    if col_idx == 7:
                        cell.fill = green_fill if num < 0 else red_fill
                    else:
                        cell.fill = green_fill if num > 0 else red_fill
                except (ValueError, TypeError):
                    pass

    for col_idx in range(1, 8):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 20

    # Stream
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"kpi_projet{project_id}_{period.year}_{period.month:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── PDF Export ────────────────────────────────────────────────────────────────

@router.get(
    "/kpis/pdf",
    summary="Export PDF — rapport KPI par projet/période/site",
    response_description="Fichier .pdf téléchargeable",
)
def export_kpis_pdf(
    project_id: int           = Query(..., description="ID du projet GitLab"),
    period_id:  Optional[int] = Query(default=None, description="ID de la période (défaut: dernière)"),
    site_id:    Optional[int] = Query(default=None, description="Filtrer par site (défaut: tous)"),
    db:         Session       = Depends(get_db),
    _:          AppUser       = Depends(get_current_user),
):
    """
    Génère un fichier PDF (.pdf) avec le rapport KPI complet en format paysage A4.

    Inclut :
    - Tableau principal des KPIs par site
    - Tableau des deltas vs période précédente
    - Résumé statistique global
    """
    try:
        from reportlab.lib                 import colors
        from reportlab.lib.pagesizes       import A4, landscape
        from reportlab.lib.styles         import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units          import cm
        from reportlab.platypus           import (
            SimpleDocTemplate, Table, TableStyle,
            Paragraph, Spacer, HRFlowable,
        )
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="reportlab non installé. Exécutez : pip install reportlab",
        )

    period, snapshots = _resolve_period_and_snapshots(db, project_id, period_id, site_id)
    period_label      = f"{MOIS_FR.get(period.month, '')} {period.year}"

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize      = landscape(A4),
        rightMargin   = 1.5 * cm,
        leftMargin    = 1.5 * cm,
        topMargin     = 1.5 * cm,
        bottomMargin  = 1.5 * cm,
    )

    styles = getSampleStyleSheet()
    BLUE   = colors.HexColor("#2563EB")
    LIGHT  = colors.HexColor("#EFF6FF")
    GREEN  = colors.HexColor("#D1FAE5")
    RED    = colors.HexColor("#FEE2E2")

    title_style = ParagraphStyle(
        "KpiTitle",
        parent    = styles["Title"],
        textColor = BLUE,
        fontSize  = 16,
        spaceAfter = 4,
    )
    subtitle_style = ParagraphStyle(
        "KpiSubtitle",
        parent    = styles["Normal"],
        textColor = colors.HexColor("#6B7280"),
        fontSize  = 10,
        spaceAfter = 10,
    )
    section_style = ParagraphStyle(
        "KpiSection",
        parent    = styles["Heading2"],
        textColor = BLUE,
        fontSize  = 12,
        spaceBefore = 14,
        spaceAfter  = 6,
    )

    story = []

    # ── Titre ─────────────────────────────────────────────────────────────────
    story.append(Paragraph(f"Rapport KPI GitLab", title_style))
    story.append(Paragraph(
        f"Projet #{project_id} &nbsp;|&nbsp; {period_label} &nbsp;|&nbsp; "
        f"Généré le {__import__('datetime').date.today().strftime('%d/%m/%Y')}",
        subtitle_style,
    ))
    story.append(HRFlowable(width="100%", thickness=1, color=BLUE, spaceAfter=10))

    # ── Statistiques globales ─────────────────────────────────────────────────
    total_commits  = sum(s.nb_commits_per_project for s in snapshots)
    total_devs     = sum(s.nb_developers         for s in snapshots)
    total_mrs      = sum(s.total_mrs_created      for s in snapshots)
    avg_approved   = (
        sum(s.approved_mr_rate for s in snapshots) / len(snapshots)
        if snapshots else 0
    )

    summary_data = [
        ["Total Commits", "Total MRs", "Total Développeurs", "Avg Approved MR Rate"],
        [
            str(total_commits),
            str(total_mrs),
            str(total_devs),
            _format_rate(avg_approved),
        ],
    ]
    summary_table = Table(summary_data, colWidths=[6 * cm] * 4)
    summary_table.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0), BLUE),
        ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, 0), 10),
        ("FONTSIZE",    (0, 1), (-1, 1), 14),
        ("FONTNAME",    (0, 1), (-1, 1), "Helvetica-Bold"),
        ("TEXTCOLOR",   (0, 1), (-1, 1), BLUE),
        ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("ROWHEIGHT",   (0, 0), (-1, -1), 28),
        ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
        ("ROUNDEDCORNERS", [4]),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 16))

    # ── Tableau KPIs principal ────────────────────────────────────────────────
    story.append(Paragraph("KPIs par site", section_style))

    kpi_headers = [
        "Site", "Devs", "MR Rate\n/Site", "Approved\nMR Rate",
        "Merged\nMR Rate", "Commit Rate\n/Site", "Commits\n/Projet",
        "Review\nMoyen (h)", "Score",
    ]
    kpi_data = [kpi_headers]

    for snap in snapshots:
        kpi_data.append([
            _get_site_name(db, snap.site_id),
            str(snap.nb_developers),
            _format_float(snap.mr_rate_per_site),
            _format_rate(snap.approved_mr_rate),
            _format_rate(snap.merged_mr_rate),
            _format_float(snap.commit_rate_per_site),
            str(snap.nb_commits_per_project),
            _format_float(snap.avg_review_time_hours, 1),
            _format_float(snap.developer_score),
        ])

    col_widths_kpi = [3.5*cm, 1.8*cm, 2.5*cm, 2.8*cm, 2.5*cm, 2.8*cm, 2.5*cm, 2.8*cm, 2.2*cm]
    kpi_table      = Table(kpi_data, colWidths=col_widths_kpi, repeatRows=1)
    kpi_style      = [
        ("BACKGROUND",  (0, 0), (-1, 0), BLUE),
        ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, -1), 9),
        ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("GRID",        (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
        ("ROWHEIGHT",   (0, 0), (-1, -1), 22),
        ("WORDWRAP",    (0, 0), (-1, -1), True),
    ]
    # Lignes alternées
    for i in range(1, len(kpi_data)):
        if i % 2 == 0:
            kpi_style.append(("BACKGROUND", (0, i), (-1, i), LIGHT))

    kpi_table.setStyle(TableStyle(kpi_style))
    story.append(kpi_table)
    story.append(Spacer(1, 16))

    # ── Tableau Deltas ────────────────────────────────────────────────────────
    story.append(Paragraph("Évolution vs période précédente", section_style))

    delta_headers = [
        "Site", "Δ MR Rate", "Δ Approved MR", "Δ Merged MR",
        "Δ Commit Rate", "Δ Commits/Projet", "Δ Review (h)",
    ]
    delta_data = [delta_headers]

    for snap in snapshots:
        delta_row = [
            _get_site_name(db, snap.site_id),
            _format_float(snap.delta_mr_rate),
            _format_rate(snap.delta_approved_mr_rate) if snap.delta_approved_mr_rate is not None else "N/A",
            _format_rate(snap.delta_merged_mr_rate)   if snap.delta_merged_mr_rate   is not None else "N/A",
            _format_float(snap.delta_commit_rate),
            str(snap.delta_nb_commits) if snap.delta_nb_commits is not None else "N/A",
            _format_float(snap.delta_avg_review_time, 1),
        ]
        delta_data.append(delta_row)

    col_widths_delta = [3.5*cm, 3*cm, 3.5*cm, 3*cm, 3.2*cm, 3.5*cm, 3*cm]
    delta_table      = Table(delta_data, colWidths=col_widths_delta, repeatRows=1)
    delta_style      = [
        ("BACKGROUND", (0, 0), (-1, 0), BLUE),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 9),
        ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("GRID",       (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
        ("ROWHEIGHT",  (0, 0), (-1, -1), 22),
    ]
    # Coloration conditionnelle des deltas
    for row_idx in range(1, len(delta_data)):
        for col_idx in range(1, len(delta_headers)):
            val = delta_data[row_idx][col_idx]
            if val not in ("N/A", None):
                try:
                    num = float(str(val).replace("%", ""))
                    # review_time : négatif = bien
                    if col_idx == 6:
                        fill = GREEN if num < 0 else RED
                    else:
                        fill = GREEN if num > 0 else RED
                    delta_style.append(
                        ("BACKGROUND", (col_idx, row_idx), (col_idx, row_idx), fill)
                    )
                except (ValueError, TypeError):
                    pass

    delta_table.setStyle(TableStyle(delta_style))
    story.append(delta_table)

    # ── Build ─────────────────────────────────────────────────────────────────
    doc.build(story)
    buf.seek(0)

    filename = f"kpi_projet{project_id}_{period.year}_{period.month:02d}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )